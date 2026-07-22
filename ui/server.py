#!/usr/bin/env python3
"""
Leave Eligibility Engine — local web UI.

Serves a single-page app for the final_pipeline leave-eligibility rule engine:
  - browse the saved eligibility rules (EE / notice / certification / leave-reason
    condition packs) for each jurisdiction
  - submit a claim (paste JSON or upload file(s)) and evaluate it against those
    rules, seeing the decision, the reasoning, and every condition that was
    checked with its pass/fail result

Usage (from anywhere — paths are auto-detected relative to this file):
    python3 ui/server.py [--port 8791] [--host 0.0.0.0] [--no-browser]

Then open http://localhost:8791 in your browser.

On SageMaker Studio / Code Editor, bind stays 0.0.0.0 (default) and open:
    https://<domain>.studio.<region>.sagemaker.aws/<app>/default/proxy/8791/
(Use jupyterlab or jupyter depending on your space; Code Editor proxy support varies.)

No third-party dependencies — stdlib only.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import webbrowser
from io import BytesIO
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

UI_DIR = Path(__file__).resolve().parent
PIPELINE_DIR = UI_DIR.parent

if str(PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_DIR))

from function.common import load_json, save_json, utc_now_iso  # noqa: E402
from function.test_case_evaluator import (  # noqa: E402
    LEAVE_REASONS,
    _condition_label,
    _gt_paths,
    _has_child_nodes,
    load_all_cases,
    run_case,
)
import run_pipeline  # noqa: E402  (reuses its jurisdiction -> chunk-file mapping)

try:
    import openpyxl  # noqa: E402
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

INDEX_HTML = UI_DIR / "index.html"
JURISDICTIONS = ["federal", "CA", "GA", "TN"]
TEMPLATES_DIR = PIPELINE_DIR / "data" / "test_cases_templates"
MANUAL_REVIEWS_PATH = PIPELINE_DIR / "output" / "manual_reviews.json"

# Columns written by ui/build_templates.py that are not claim facts.
_NON_FACT_COLUMNS = {"claim_id", "jurisdiction", "leave_reason", "narrative"}

# Evaluation source: rules are always evaluated against the saved condition
# packs (internally keyed "gt" in the evaluator module — not surfaced in the UI).
SOURCE = "gt"

# SageMaker / Jupyter proxies forward URLs like /proxy/8791/api/rules (or
# /codeeditor/default/proxy/8791/...). Strip that prefix so route matching works.
_PROXY_PREFIX_RE = re.compile(r"^(?:/[^/]+)*/proxy/\d+")


def _route_path(raw: str) -> str:
    """Normalize request path: drop query string and any /proxy/<port> prefix."""
    path = urlparse(raw).path or "/"
    stripped = _PROXY_PREFIX_RE.sub("", path)
    if not stripped or stripped == "/":
        return "/"
    return stripped if stripped.startswith("/") else f"/{stripped}"


def _rel(p: Path) -> str:
    try:
        return str(p.relative_to(PIPELINE_DIR))
    except ValueError:
        return str(p)


def _annotate_node(node: Any) -> Any:
    """Recursively attach a human-readable `label` to a condition node (and its
    children), reusing the same de-slugging logic the evaluator's citation
    checklist uses, so the Rules browser and the Evaluate checklist describe
    conditions the same way instead of the raw extracted field/operator form."""
    if not isinstance(node, dict):
        return node
    node = dict(node)
    node["label"] = _condition_label(node)
    rv = node.get("required_value")
    if _has_child_nodes(rv):
        node["required_value"] = [_annotate_node(c) for c in rv]
    return node


def _annotate_pack(data: Any) -> Any:
    if not isinstance(data, dict):
        return data
    data = dict(data)
    if isinstance(data.get("conditions"), list):
        data["conditions"] = [_annotate_node(c) for c in data["conditions"]]
    if isinstance(data.get("employee_eligibility_conditions"), list):
        data["employee_eligibility_conditions"] = [
            _annotate_node(c) for c in data["employee_eligibility_conditions"]
        ]
    if isinstance(data.get("included_relationships"), dict):
        rel = dict(data["included_relationships"])
        rel["label"] = "Included relationships"
        data["included_relationships"] = rel
    if "condition_name" in data or "field" in data or "operator" in data:
        data = _annotate_node(data)
    return data


def _load_rules() -> Dict[str, Any]:
    """Build a browsable structure of every saved condition pack per jurisdiction."""
    rules: Dict[str, Any] = {}
    for j in JURISDICTIONS:
        packs: Dict[str, Any] = {"leave_reasons": {}}
        base_paths = _gt_paths(j, "health_condition")  # ee/notice/cert stable across reasons
        for key in ("ee", "notice", "cert"):
            p = base_paths[key]
            if p.exists():
                packs[key] = {"path": _rel(p), "data": _annotate_pack(load_json(p))}
        for reason in sorted(LEAVE_REASONS):
            paths = _gt_paths(j, reason)
            p = paths["leave_reason"]
            if p.exists():
                packs["leave_reasons"][reason] = {"path": _rel(p), "data": _annotate_pack(load_json(p))}
        rules[j] = packs
    return rules


def _load_chunks() -> Dict[str, List[Dict[str, Any]]]:
    """Raw legal-source chunks per jurisdiction, keyed the same way regulation
    citations appear in the rule packs, for the Rules/Evaluate 'view source
    text' features."""
    out: Dict[str, List[Dict[str, Any]]] = {}
    for j in JURISDICTIONS:
        cfg = run_pipeline.JURISDICTIONS.get(j, {})
        chunks: List[Dict[str, Any]] = []
        for fname in cfg.get("chunk_files", []):
            p = PIPELINE_DIR / "data" / j / fname
            if not p.exists():
                continue
            data = load_json(p)
            if isinstance(data, list):
                chunks.extend(data)
        out[j] = [
            {
                "regulation_number": c.get("regulation number", []),
                "function": c.get("function", []),
                "reason": c.get("reason", []),
                "raw_text": c.get("raw text", ""),
            }
            for c in chunks
        ]
    return out


def _append_manual_review(payload: Dict[str, Any]) -> Dict[str, Any]:
    required = ["claim_id", "jurisdiction", "leave_reason", "manual_decision"]
    missing = [k for k in required if not payload.get(k)]
    if missing:
        raise ValueError(f"Missing field(s): {', '.join(missing)}")
    if payload["manual_decision"] not in ("APPROVED", "DENIED"):
        raise ValueError("manual_decision must be 'APPROVED' or 'DENIED'")

    existing: List[Dict[str, Any]] = []
    if MANUAL_REVIEWS_PATH.exists():
        try:
            loaded = load_json(MANUAL_REVIEWS_PATH)
            existing = loaded if isinstance(loaded, list) else []
        except Exception:  # noqa: BLE001
            existing = []

    record = {
        "claim_id": payload["claim_id"],
        "jurisdiction": payload["jurisdiction"],
        "leave_reason": payload["leave_reason"],
        "engine_decision": payload.get("engine_decision"),
        "manual_decision": payload["manual_decision"],
        "reviewer_note": payload.get("reviewer_note", ""),
        "timestamp": utc_now_iso(),
    }
    existing.append(record)
    save_json(existing, MANUAL_REVIEWS_PATH)
    return record


def _examples() -> Dict[str, Any]:
    """One representative sample claim per jurisdiction + leave-reason combo."""
    cases = load_all_cases()
    out: Dict[str, Any] = {}
    for c in cases:
        key = f"{c.get('jurisdiction')}::{c.get('leave_reason')}"
        prefer = c.get("case_type") == "pass"
        if key not in out or (prefer and out[key].get("_case_type") != "pass"):
            out[key] = {
                "jurisdiction": c.get("jurisdiction"),
                "leave_reason": c.get("leave_reason"),
                "narrative": c.get("narrative", ""),
                "facts": c.get("facts", {}),
                "_case_type": c.get("case_type"),
            }
    for v in out.values():
        v.pop("_case_type", None)
    return out


def _parse_multipart(content_type: str, body: bytes) -> List[Dict[str, Any]]:
    """Minimal RFC 7578 multipart/form-data parser (stdlib only). Returns the
    file parts as [{"field": name, "filename": ..., "content": bytes}, ...]."""
    m = re.search(r'boundary="?([^";]+)"?', content_type or "")
    if not m:
        return []
    boundary = ("--" + m.group(1)).encode()
    files: List[Dict[str, Any]] = []
    for raw_part in body.split(boundary):
        part = raw_part
        if part in (b"", b"--") or part.startswith(b"--"):
            continue
        if part.startswith(b"\r\n"):
            part = part[2:]
        if b"\r\n\r\n" not in part:
            continue
        header_blob, content = part.split(b"\r\n\r\n", 1)
        if content.endswith(b"\r\n"):
            content = content[:-2]
        headers_text = header_blob.decode("utf-8", errors="replace")
        fn_match = re.search(r'filename="([^"]*)"', headers_text)
        name_match = re.search(r'name="([^"]*)"', headers_text)
        if fn_match and fn_match.group(1):
            files.append({
                "field": name_match.group(1) if name_match else None,
                "filename": fn_match.group(1),
                "content": content,
            })
    return files


def _coerce_cell(value: Any) -> Any:
    if isinstance(value, str):
        v = value.strip()
        low = v.lower()
        if low in ("true", "yes"):
            return True
        if low in ("false", "no"):
            return False
        return v
    return value


def _rows_from_sheet(ws: Any) -> List[Dict[str, Any]]:
    """Extract claim rows from one worksheet built by ui/build_templates.py:
    a header row starting with 'claim_id', followed by data rows (example
    rows and/or user-entered rows), skipping the '↓ enter your own claims
    below ↓' separator and any untouched blank template rows."""
    header_row_idx: Optional[int] = None
    for r in range(1, ws.max_row + 1):
        first_cell = ws.cell(row=r, column=1).value
        if isinstance(first_cell, str) and first_cell.strip().lower() == "claim_id":
            header_row_idx = r
            break
    if header_row_idx is None:
        return []

    headers: List[Optional[str]] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row_idx, column=c).value
        headers.append(str(v).strip() if v is not None else None)

    claims: List[Dict[str, Any]] = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c + 1).value for c in range(len(headers))]
        first = row_vals[0] if row_vals else None
        if isinstance(first, str) and first.strip().startswith("↓"):
            continue
        row_map = {headers[i]: _coerce_cell(row_vals[i]) for i in range(len(headers)) if headers[i]}
        facts = {k: v for k, v in row_map.items() if k not in _NON_FACT_COLUMNS and v not in (None, "")}
        has_content = bool(facts) or row_map.get("claim_id") or row_map.get("narrative")
        if not has_content:
            continue
        claims.append({
            "id": row_map.get("claim_id") or None,
            "jurisdiction": row_map.get("jurisdiction"),
            "leave_reason": row_map.get("leave_reason"),
            "narrative": row_map.get("narrative") or "",
            "facts": facts,
        })
    return claims


def _claims_from_xlsx_bytes(data: bytes) -> List[Dict[str, Any]]:
    if openpyxl is None:
        raise RuntimeError(
            "The server needs the 'openpyxl' package to read Excel uploads. "
            "Install it with: pip install openpyxl"
        )
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    claims: List[Dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == "instructions":
            continue
        claims.extend(_rows_from_sheet(wb[sheet_name]))
    return claims


def _normalize_claim(raw: Any, idx: int) -> Dict[str, Any]:
    if not isinstance(raw, dict):
        raise ValueError(f"Claim #{idx + 1}: expected a JSON object")

    if isinstance(raw.get("facts"), dict):
        case = dict(raw)
        facts = dict(raw["facts"])
    else:
        facts = dict(raw)
        case = {}

    jurisdiction = case.get("jurisdiction") or facts.get("jurisdiction")
    leave_reason = case.get("leave_reason") or facts.get("leave_reason")

    if not jurisdiction:
        raise ValueError(f"Claim #{idx + 1}: missing 'jurisdiction'")
    if not leave_reason:
        raise ValueError(f"Claim #{idx + 1}: missing 'leave_reason'")

    case["id"] = case.get("id") or raw.get("id") or f"claim-{idx + 1}"
    case["jurisdiction"] = jurisdiction
    case["leave_reason"] = leave_reason
    case["facts"] = facts
    case.setdefault("label", raw.get("label", ""))
    case.setdefault("narrative", raw.get("narrative", ""))
    return case


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):  # quieter console
        pass

    def _send_json(self, obj, status=200):
        body = json.dumps(obj).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def _send_html(self, path: Path):
        if not path.exists():
            self.send_error(404, "Not found")
            return
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, path: Path, download_name: Optional[str] = None):
        if not path.exists():
            self.send_error(404, "Not found")
            return
        body = path.read_bytes()
        self.send_response(200)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Length", str(len(body)))
        self.send_header(
            "Content-Disposition", f'attachment; filename="{download_name or path.name}"'
        )
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        path = _route_path(self.path)
        if path == "/" or path == "/index.html":
            self._send_html(INDEX_HTML)
        elif path == "/api/rules":
            try:
                self._send_json(_load_rules())
            except Exception as e:  # noqa: BLE001
                self._send_json({"error": str(e)}, status=500)
        elif path == "/api/examples":
            try:
                self._send_json(_examples())
            except Exception as e:  # noqa: BLE001
                self._send_json({"error": str(e)}, status=500)
        elif path == "/api/chunks":
            try:
                self._send_json(_load_chunks())
            except Exception as e:  # noqa: BLE001
                self._send_json({"error": str(e)}, status=500)
        elif path.startswith("/templates/"):
            name = path[len("/templates/"):].strip("/")
            juris = name[:-5] if name.endswith(".xlsx") else name
            if juris not in JURISDICTIONS:
                self.send_error(404, f"No template for '{juris}'")
                return
            self._send_file(TEMPLATES_DIR / f"{juris}.xlsx", download_name=f"{juris}_claim_template.xlsx")
        else:
            self.send_error(404, f"Not found: {path}")

    def do_POST(self):
        path = _route_path(self.path)
        if path == "/api/evaluate":
            try:
                length = int(self.headers.get("Content-Length", 0) or 0)
                raw = self.rfile.read(length) if length else b"{}"
                payload = json.loads(raw or b"{}")
                claims = payload.get("claims") or payload.get("cases") or []
                if not isinstance(claims, list) or not claims:
                    raise ValueError("No claims submitted")
                normalized = [_normalize_claim(c, i) for i, c in enumerate(claims)]
                results: List[Dict[str, Any]] = [run_case(c, SOURCE) for c in normalized]
                self._send_json({"results": results})
            except Exception as e:  # noqa: BLE001
                self._send_json({"error": str(e)}, status=400)
        elif path == "/api/evaluate-upload":
            try:
                length = int(self.headers.get("Content-Length", 0) or 0)
                body = self.rfile.read(length) if length else b""
                files = _parse_multipart(self.headers.get("Content-Type", ""), body)
                if not files:
                    raise ValueError("No file(s) received")
                claims: List[Dict[str, Any]] = []
                for f in files:
                    if not f["filename"].lower().endswith(".xlsx"):
                        raise ValueError(f"'{f['filename']}' is not an .xlsx file")
                    claims.extend(_claims_from_xlsx_bytes(f["content"]))
                if not claims:
                    raise ValueError("No claim rows found in the uploaded file(s)")
                normalized = [_normalize_claim(c, i) for i, c in enumerate(claims)]
                results: List[Dict[str, Any]] = [run_case(c, SOURCE) for c in normalized]
                self._send_json({"results": results})
            except Exception as e:  # noqa: BLE001
                self._send_json({"error": str(e)}, status=400)
        elif path == "/api/manual-decision":
            try:
                length = int(self.headers.get("Content-Length", 0) or 0)
                raw = self.rfile.read(length) if length else b"{}"
                payload = json.loads(raw or b"{}")
                record = _append_manual_review(payload)
                self._send_json({"ok": True, "record": record})
            except Exception as e:  # noqa: BLE001
                self._send_json({"error": str(e)}, status=400)
        else:
            self.send_error(404, f"Not found: {path}")

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=8791)
    # 0.0.0.0 so SageMaker / remote proxies can reach the process (localhost-only
    # binds are unreachable from the Studio proxy sidecar).
    parser.add_argument("--host", default="0.0.0.0")
    parser.add_argument("--no-browser", action="store_true")
    args = parser.parse_args()

    server = ThreadingHTTPServer((args.host, args.port), Handler)
    local_url = f"http://127.0.0.1:{args.port}"
    print(f"Leave Eligibility Engine listening on {args.host}:{args.port}  (Ctrl+C to stop)")
    print(f"  Local:     {local_url}")
    print(f"  SageMaker: open …/proxy/{args.port}/ on your Studio / Code Editor URL")
    print(f"             (e.g. https://<domain>.studio.<region>.sagemaker.aws/jupyterlab/default/proxy/{args.port}/)")
    if not args.no_browser:
        try:
            webbrowser.open(local_url)
        except Exception:  # noqa: BLE001
            pass
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
