#!/usr/bin/env python3
"""Generate per-jurisdiction Excel claim-input templates from the existing
synthetic test cases (data/test_cases/*/*.json), one sheet per leave reason,
with the 3 existing sample claims as example rows plus blank rows to fill in.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

PIPELINE_DIR = Path("/sessions/wonderful-dreamy-fermi/mnt/Final_pipeline")
TEST_CASES_DIR = PIPELINE_DIR / "data" / "test_cases"
OUT_DIR = PIPELINE_DIR / "data" / "test_cases_templates"
OUT_DIR.mkdir(parents=True, exist_ok=True)

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="123152")
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=10)
LEGEND_FONT = Font(name=FONT_NAME, italic=True, size=9, color="555555")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13, color="123152")
EXAMPLE_FILL = PatternFill("solid", fgColor="EAF2FD")
EXAMPLE_FONT = Font(name=FONT_NAME, size=10, color="1F4C7D")
BLANK_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="D9DEE6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FIXED_COLUMNS = ["claim_id", "jurisdiction", "leave_reason", "narrative"]
BLANK_ROWS_TO_ADD = 12


def sheet_name_for(reason: str) -> str:
    return reason.replace("_", " ").title()[:31]


def build_sheet(ws: Worksheet, jurisdiction: str, reason: str, cases: list[dict]) -> None:
    fact_keys: list[str] = []
    for c in cases:
        for k in c.get("facts", {}).keys():
            if k.startswith("_") or k in FIXED_COLUMNS:
                continue
            if k not in fact_keys:
                fact_keys.append(k)

    columns = FIXED_COLUMNS + fact_keys
    n_cols = len(columns)

    # --- Title + legend (rows 1-3), header (row 5), data starting row 6 ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=f"Claim input template — {jurisdiction} / {reason.replace('_',' ').title()}")
    title_cell.font = TITLE_FONT

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=n_cols)
    legend = ws.cell(
        row=2, column=1,
        value=(
            "How to use: rows shaded blue below are example claims from the rule set — do not edit them, "
            "copy a row instead. Add your own claims starting after the blank separator row. Keep 'jurisdiction' "
            "and 'leave_reason' as shown (they select which saved rules the claim is checked against). "
            "Enter TRUE / FALSE for yes-no facts, plain numbers for counts/durations, and text for everything else. "
            "Leave a cell blank if that fact does not apply to your claim — 'claim_id' is optional (one is assigned automatically if left blank)."
        ),
    )
    legend.font = LEGEND_FONT
    legend.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 46

    header_row = 5
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # --- Example rows (from the real synthetic cases) ---
    row = header_row + 1
    for c in cases:
        facts = c.get("facts", {})
        values = {
            "claim_id": c.get("id", ""),
            "jurisdiction": c.get("jurisdiction", jurisdiction),
            "leave_reason": c.get("leave_reason", reason),
            "narrative": c.get("narrative", ""),
        }
        for k in fact_keys:
            values[k] = facts.get(k, "")
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=values.get(col_name, ""))
            cell.font = EXAMPLE_FONT
            cell.fill = EXAMPLE_FILL
            cell.border = BORDER
        row += 1

    # --- Blank separator row, labeled ---
    ws.cell(row=row, column=1, value="↓ enter your own claims below ↓").font = LEGEND_FONT
    row += 1

    # --- Blank rows for user entry, pre-filled with jurisdiction/leave_reason ---
    for _ in range(BLANK_ROWS_TO_ADD):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = BLANK_FONT
            cell.border = BORDER
            if col_name == "jurisdiction":
                cell.value = jurisdiction
            elif col_name == "leave_reason":
                cell.value = reason
        row += 1

    # Column widths
    for col_idx, col_name in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        width = max(14, min(34, len(col_name) + 4))
        ws.column_dimensions[letter].width = width


def build_workbook(jurisdiction: str, pairs: list[dict]) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # Table-of-contents sheet
    toc = wb.create_sheet("Instructions", 0)
    toc.sheet_view.showGridLines = False
    toc.merge_cells("A1:C1")
    c = toc.cell(row=1, column=1, value=f"Claim input templates — {jurisdiction}")
    c.font = TITLE_FONT
    toc.merge_cells("A2:C4")
    instr = toc.cell(
        row=2, column=1,
        value=(
            "Each tab below is a leave-reason pathway available in this jurisdiction. Open the matching tab, "
            "add one row per claim using the example rows as a guide, save the file, then upload it in the "
            "Evaluate screen of the Leave Eligibility Engine. You can fill in as many or as few tabs as you need."
        ),
    )
    instr.font = LEGEND_FONT
    instr.alignment = Alignment(wrap_text=True, vertical="top")
    toc.row_dimensions[2].height = 60
    toc.column_dimensions["A"].width = 4
    toc.column_dimensions["B"].width = 40
    toc.column_dimensions["C"].width = 14

    hdr_row = 6
    for i, label in enumerate(["#", "Leave reason tab", "Sample claims"]):
        cell = toc.cell(row=hdr_row, column=1 + i, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for i, p in enumerate(pairs, start=1):
        r = hdr_row + i
        toc.cell(row=r, column=1, value=i).border = BORDER
        name_cell = toc.cell(row=r, column=2, value=sheet_name_for(p["leave_reason"]))
        name_cell.border = BORDER
        name_cell.font = Font(name=FONT_NAME, size=10)
        toc.cell(row=r, column=3, value=p["case_count"]).border = BORDER

    for p in pairs:
        cases = json.load(open(TEST_CASES_DIR / p["path"]))
        ws = wb.create_sheet(sheet_name_for(p["leave_reason"]))
        ws.sheet_view.showGridLines = False
        build_sheet(ws, jurisdiction, p["leave_reason"], cases)

    out_path = OUT_DIR / f"{jurisdiction}.xlsx"
    wb.save(out_path)
    return out_path


def main() -> None:
    index = json.load(open(TEST_CASES_DIR / "index.json"))
    by_jurisdiction: dict[str, list[dict]] = {}
    for p in index["pairs"]:
        by_jurisdiction.setdefault(p["jurisdiction"], []).append(p)

    for jurisdiction, pairs in by_jurisdiction.items():
        path = build_workbook(jurisdiction, pairs)
        print(f"wrote {path}")


if __name__ == "__main__":
    main()
